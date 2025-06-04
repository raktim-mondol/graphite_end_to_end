#!/usr/bin/env python3
"""
HierGAT-specific visualization main file.

This script provides comprehensive HierGAT visualization showing:
- Individual Level visualizations (Level 0, Level 1, Level 2)
- Multilevel fusion with user-configurable weights
- Core mask and annotated mask visualization
- Performance metrics evaluation for all levels
- Excel export of metrics for each core
- All components in a single comprehensive figure

Usage examples:
    # Default weights (0.5, 0.3, 0.2) with metrics
    python main_multi_level_fusion.py
    
    # Custom weights with metrics
    python main_multi_level_fusion.py --level_weights 0.4 0.4 0.2
    
    # Skip metrics calculation
    python main_multi_level_fusion.py --calculate_metrics False
    
    # Custom metrics thresholds
    python main_multi_level_fusion.py --metrics_thresholds 0.2 0.4 0.6 0.8
    
    # Process single image with metrics
    python main_multi_level_fusion.py --single_image path/to/image.png
    
    # Custom output directory
    python mmain_multi_level_fusion.py --output_suffix custom_run
"""

import random
import numpy as np
import torch
import os

def set_seed(seed=25):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    os.environ["PYTHONHASHSEED"] = str(seed)

set_seed()

from utils.imports import *
import torch
import os
from pathlib import Path
from models.hiergat import HierGATSSL
from data.slide_processor import CustomSlide
from utils.model_utils import load_mil_model
from data.dataset import CoreImageProcessor
from models.graph_builder import HierarchicalGraphBuilder
import logging
import json
import cv2
import random
import numpy as np
import matplotlib.pyplot as plt
from scipy.ndimage import gaussian_filter
from data.patch_extractor import MultiscalePatchExtractor
from models.graph_builder import HierarchicalGraphBuilder
from models.hiergat import HierGATSSL
from training.trainer import HierGATSSLTrainer
from models.hiergat import HierGATSSL
import argparse
from models.graph_builder import HierarchicalGraphBuilder
from data.dataset import CoreImageProcessor, HierGATSSLDataset
from training.losses import HierarchicalInfoMaxLoss, LossTracker
from torch_geometric.loader import DataLoader as PyGDataLoader
from models.inference import HierGATSSLInference
from utils.extract_core_mask import CoreExtractor
from PIL import Image
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


def calculate_iou(pred_mask, true_mask):
    """Calculate Intersection over Union between two binary masks."""
    intersection = np.logical_and(pred_mask, true_mask)
    union = np.logical_or(pred_mask, true_mask)
    return np.sum(intersection) / np.sum(union) if np.sum(union) > 0 else 0


def calculate_precision(pred_mask, true_mask):
    """Calculate precision between two binary masks."""
    true_positive = np.sum(np.logical_and(pred_mask, true_mask))
    false_positive = np.sum(np.logical_and(pred_mask, np.logical_not(true_mask)))
    return true_positive / (true_positive + false_positive) if (true_positive + false_positive) > 0 else 0


def calculate_recall(pred_mask, true_mask):
    """Calculate recall between two binary masks."""
    true_positive = np.sum(np.logical_and(pred_mask, true_mask))
    false_negative = np.sum(np.logical_and(np.logical_not(pred_mask), true_mask))
    return true_positive / (true_positive + false_negative) if (true_positive + false_negative) > 0 else 0


def calculate_f1_score(precision, recall):
    """Calculate F1 score from precision and recall."""
    return 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0


def calculate_comprehensive_metrics(pred_attention, gt_mask, thresholds=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]):
    """
    Calculate comprehensive metrics for attention maps vs ground truth masks.
    
    Args:
        pred_attention: Predicted attention map (continuous values 0-1)
        gt_mask: Ground truth binary mask
        thresholds: List of thresholds to convert attention to binary mask
        
    Returns:
        Dictionary containing metrics for each threshold
    """
    # Normalize ground truth to binary
    gt_binary = (gt_mask > 0).astype(int)
    
    # Convert RGB to grayscale if needed
    if len(gt_binary.shape) == 3:
        gt_binary = np.mean(gt_binary, axis=2)
        gt_binary = (gt_binary > 0.5).astype(int)
    
    results = {}
    
    for threshold in thresholds:
        # Convert attention to binary mask
        pred_binary = (pred_attention > threshold).astype(int)
        
        # Calculate confusion matrix elements
        TP = np.sum((pred_binary == 1) & (gt_binary == 1))
        FP = np.sum((pred_binary == 1) & (gt_binary == 0))
        FN = np.sum((pred_binary == 0) & (gt_binary == 1))
        TN = np.sum((pred_binary == 0) & (gt_binary == 0))
        
        # Calculate metrics
        precision = TP / (TP + FP) if (TP + FP) > 0 else 0
        recall = TP / (TP + FN) if (TP + FN) > 0 else 0
        f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        iou = TP / (TP + FP + FN) if (TP + FP + FN) > 0 else 0
        accuracy = (TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) > 0 else 0
        specificity = TN / (TN + FP) if (TN + FP) > 0 else 0
        
        results[f'threshold_{threshold}'] = {
            'threshold': threshold,
            'precision': precision,
            'recall': recall,
            'f1_score': f1_score,
            'iou': iou,
            'accuracy': accuracy,
            'specificity': specificity,
            'TP': TP,
            'FP': FP,
            'FN': FN,
            'TN': TN
        }
    
    return results


def create_metrics_excel(image_name, level_metrics, save_dir):
    """
    Create Excel file with metrics for all levels and multilevel fusion.
    
    Args:
        image_name: Name of the image/core
        level_metrics: Dictionary containing metrics for each level
        save_dir: Directory to save the Excel file
    """
    try:
        save_dir = Path(save_dir)
        save_dir.mkdir(parents=True, exist_ok=True)
        
        # Create Excel file
        excel_path = save_dir / f'metrics_{Path(image_name).stem}.xlsx'
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
            # Define the order of levels
            level_order = ['Level_0', 'Level_1', 'Level_2', 'Multilevel_Fusion']
            
            for level_name in level_order:
                if level_name in level_metrics:
                    metrics_data = level_metrics[level_name]
                    
                    # Convert to DataFrame
                    df_data = []
                    for threshold_key, metrics in metrics_data.items():
                        df_data.append(metrics)
                    
                    df = pd.DataFrame(df_data)
                    
                    # Round numerical values for better display
                    numeric_cols = ['precision', 'recall', 'f1_score', 'iou', 'accuracy', 'specificity']
                    for col in numeric_cols:
                        if col in df.columns:
                            df[col] = df[col].round(4)
                    
                    # Write to Excel sheet
                    df.to_excel(writer, sheet_name=level_name, index=False)
                    
                    # Get the workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets[level_name]
                    
                    # Format headers
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 15)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Highlight best performing thresholds
                    if len(df) > 0:
                        # Find best F1 score row
                        best_f1_idx = df['f1_score'].idxmax() + 2  # +2 for 0-indexing and header
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=best_f1_idx, column=col_idx)
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        
                        # Find best IoU score row
                        best_iou_idx = df['iou'].idxmax() + 2
                        if best_iou_idx != best_f1_idx:
                            for col_idx in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=best_iou_idx, column=col_idx)
                                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        
        print(f"✅ Metrics Excel file saved: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {str(e)}")
        traceback.print_exc()
        return None


def hiergat_multilevel_fusion(smooth_maps, level_weights=[0.5, 0.3, 0.2]):
    """
    Generate fused HierGAT map from multiple levels with custom weights
    
    Args:
        smooth_maps: Dictionary containing smoothed maps for each level
        level_weights: List of weights for [level_0, level_1, level_2]
    """
    try:
        # Normalize weights
        level_weights = np.array(level_weights)
        level_weights = level_weights / level_weights.sum()
        
        # Get shape from first map
        H, W = next(iter(smooth_maps.values())).shape
        fused_map = np.zeros((H, W))
        
        # Combine levels with weighted fusion
        for level in range(3):
            if level in smooth_maps:
                level_map = smooth_maps[level].copy()
                # Normalize level map
                level_map = (level_map - level_map.min()) / (level_map.max() - level_map.min() + 1e-8)
                # Apply additional smoothing based on level
                sigma = 150 * (2 ** level)
                smoothed = gaussian_filter(level_map, sigma=sigma)
                # Add to fusion with weight
                fused_map += level_weights[level] * smoothed
        
        # Final normalization
        fused_map = (fused_map - fused_map.min()) / (fused_map.max() - fused_map.min() + 1e-8)
        
        return fused_map, level_weights
        
    except Exception as e:
        print(f"Error in multilevel fusion: {str(e)}")
        traceback.print_exc()
        raise


def hiergat_comprehensive_visualization(gat_results, image_path, save_dir=None, mask_dir=None, level_weights=[0.5, 0.3, 0.2], calculate_metrics=True, metrics_thresholds=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]):
    """
    Comprehensive HierGAT visualization showing all levels and fusion with performance metrics
    
    Args:
        gat_results: Results from GAT inference
        image_path: Path to the image
        save_dir: Directory to save visualization
        mask_dir: Directory containing mask files
        level_weights: Weights for multilevel fusion [level_0, level_1, level_2]
        calculate_metrics: Boolean to calculate performance metrics
        metrics_thresholds: List of thresholds for metrics calculation
    """
    try:
        if save_dir:
            save_dir = Path(save_dir)
            save_dir.mkdir(parents=True, exist_ok=True)
            
        # Initialize core extractor
        core_extractor = CoreExtractor(threshold=200, kernel_size=7)
            
        # Read original image and get core mask
        original = cv2.imread(str(image_path))
        original = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
        H, W = original.shape[:2]
        
        # Extract core mask and contours
        core_mask = core_extractor.extract_core_mask(original)
        core_contours = core_extractor.get_core_boundary(core_mask)
        
        # Initialize attention maps for each level
        gat_maps = {level: np.zeros((H, W)) for level in range(3)}
        
        # Process GAT levels deterministically
        viz_data = gat_results['viz_data']
        for level in range(3):
            level_patches = [p for p in viz_data['patches']['kept'] if p['level'] == level]
            level_patches.sort(key=lambda x: (x['level_coords'][1], x['level_coords'][0]))
            
            if not level_patches:
                continue
                
            level_attention = gat_results['attention_weights'][f'level_{level}'].cpu().numpy().squeeze()
            level_attention = (level_attention - level_attention.min()) / (level_attention.max() - level_attention.min() + 1e-8)
            
            scale_factor = 2 ** level
            for idx, patch in enumerate(level_patches):
                if idx >= len(level_attention):
                    break
                    
                base_x, base_y = patch['base_coords']
                patch_size = patch['size'][0] * scale_factor
                
                y_end = min(base_y + patch_size, H)
                x_end = min(base_x + patch_size, W)
                
                if y_end > base_y and x_end > base_x:
                    gat_maps[level][base_y:y_end, base_x:x_end] = level_attention[idx]
        
        # Apply core masks to all levels
        for level in gat_maps:
            gat_maps[level] = core_extractor.apply_core_mask(gat_maps[level], core_mask)
        
        # Apply normalization and smoothing for each level
        smooth_maps = {}
        for level in range(3):
            # First normalize
            level_map = gat_maps[level].copy()
            level_map = (level_map - level_map.min()) / (level_map.max() - level_map.min() + 1e-8)
           
            # Then apply smoothing
            sigma = 150 * (2 ** level)
            smoothed = gaussian_filter(level_map, sigma=sigma)
           
            # Apply core mask and store
            smooth_maps[level] = core_extractor.apply_core_mask(smoothed, core_mask)

        # Generate multilevel fusion with custom weights
        multilevel_fusion, actual_weights = hiergat_multilevel_fusion(smooth_maps, level_weights)
        
        # Load corresponding annotated mask
        image_name = Path(image_path).name
        if mask_dir:
            mask_path = Path(mask_dir) / image_name
        else:
            mask_path = Path("../../dataset/training_dataset_step_2/mask") / image_name
       
        annotated_mask = None
        level_metrics = {}
        
        if mask_path.exists():
            annotated_mask = cv2.imread(str(mask_path))
            annotated_mask = cv2.cvtColor(annotated_mask, cv2.COLOR_BGR2RGB)
            
            if calculate_metrics:
                print(f"📊 Calculating performance metrics for {image_name}...")
                
                # Calculate metrics for each level
                for level in range(3):
                    if level in smooth_maps:
                        print(f"   📈 Calculating metrics for Level {level}")
                        level_metrics[f'Level_{level}'] = calculate_comprehensive_metrics(
                            smooth_maps[level], annotated_mask, thresholds=metrics_thresholds
                        )
                
                # Calculate metrics for multilevel fusion
                print(f"   📈 Calculating metrics for Multilevel Fusion")
                level_metrics['Multilevel_Fusion'] = calculate_comprehensive_metrics(
                    multilevel_fusion, annotated_mask, thresholds=metrics_thresholds
                )
                
                # Create Excel file with metrics
                if save_dir:
                    excel_path = create_metrics_excel(image_name, level_metrics, save_dir)
            else:
                print(f"📊 Skipping metrics calculation for {image_name}")
            
        else:
            print(f"Warning: Annotated mask not found for {image_name}")
            annotated_mask = np.zeros_like(original)

        # Create comprehensive visualization
        plt.figure(figsize=(20, 12))

        # Row 1: Original Image, Core Mask, Annotated Mask
        plt.subplot(3, 6, 1)
        plt.imshow(original)
        plt.title('Original Image', fontsize=14, fontweight='bold')
        plt.axis('off')

        plt.subplot(3, 6, 2)
        plt.imshow(original)
        for contour in core_contours:
            plt.plot(contour[:, 0, 0], contour[:, 0, 1], 'r-', linewidth=2)
        plt.title('Core Mask Overlay', fontsize=14, fontweight='bold')
        plt.axis('off')

        plt.subplot(3, 6, 3)
        if annotated_mask is not None:
            plt.imshow(annotated_mask)
        else:
            plt.imshow(np.zeros_like(original))
        plt.title('Annotated Mask', fontsize=14, fontweight='bold')
        plt.axis('off')

        # Empty spaces for alignment
        plt.subplot(3, 6, 4)
        plt.axis('off')
        plt.subplot(3, 6, 5)
        plt.axis('off')
        plt.subplot(3, 6, 6)
        plt.axis('off')

        # Row 2: Individual GAT Levels (0, 1, 2) with weights info
        for level in range(3):
            plt.subplot(3, 6, 7 + level)
            im = plt.imshow(smooth_maps[level], cmap='jet', vmin=0, vmax=1)
            
            # Add metrics info to title if available
            title = f'HierGAT Level {level}\n(Weight: {actual_weights[level]:.3f})'
            if f'Level_{level}' in level_metrics:
                # Get best F1 score for display
                best_f1 = max([m['f1_score'] for m in level_metrics[f'Level_{level}'].values()])
                best_iou = max([m['iou'] for m in level_metrics[f'Level_{level}'].values()])
                title += f'\nF1: {best_f1:.3f}, IoU: {best_iou:.3f}'
            
            plt.title(title, fontsize=12, fontweight='bold')
            plt.colorbar(im, fraction=0.046, pad=0.04)
            plt.axis('off')

        # Multilevel Fusion
        plt.subplot(3, 6, 10)
        im = plt.imshow(multilevel_fusion, cmap='jet', vmin=0, vmax=1)
        title = 'Multilevel Fusion\n(Weighted Combination)'
        if 'Multilevel_Fusion' in level_metrics:
            # Get best F1 score for display
            best_f1 = max([m['f1_score'] for m in level_metrics['Multilevel_Fusion'].values()])
            best_iou = max([m['iou'] for m in level_metrics['Multilevel_Fusion'].values()])
            title += f'\nF1: {best_f1:.3f}, IoU: {best_iou:.3f}'
        
        plt.title(title, fontsize=12, fontweight='bold')
        plt.colorbar(im, fraction=0.046, pad=0.04)
        plt.axis('off')

        # Empty spaces
        plt.subplot(3, 6, 11)
        plt.axis('off')
        plt.subplot(3, 6, 12)
        plt.axis('off')

        # Row 3: Detailed visualizations
        # Core mask only
        plt.subplot(3, 6, 13)
        plt.imshow(core_mask, cmap='gray')
        plt.title('Extracted Core Mask', fontsize=12, fontweight='bold')
        plt.axis('off')

        # Fusion overlay on original
        plt.subplot(3, 6, 14)
        plt.imshow(original)
        attention_threshold = 0.1
        attention_mask = np.ma.masked_array(
            multilevel_fusion,
            mask=~((multilevel_fusion > attention_threshold) & (core_mask > 0))
        )
        im = plt.imshow(attention_mask, cmap='jet', alpha=0.6, vmin=0, vmax=1)
        plt.title('Fusion Overlay\n(on Original)', fontsize=12, fontweight='bold')
        plt.colorbar(im, fraction=0.046, pad=0.04)
        plt.axis('off')

        # High attention regions
        plt.subplot(3, 6, 15)
        high_attention = multilevel_fusion > np.percentile(multilevel_fusion, 80)
        high_attention_masked = high_attention & (core_mask > 0)
        plt.imshow(original)
        plt.imshow(high_attention_masked, cmap='Reds', alpha=0.7)
        plt.title('High Attention\nRegions (>80th %ile)', fontsize=12, fontweight='bold')
        plt.axis('off')

        # Statistics and weights info
        plt.subplot(3, 6, 16)
        plt.axis('off')
        
        # Calculate statistics
        stats_text = f"""HierGAT Analysis Statistics:

Level Weights:
• Level 0: {actual_weights[0]:.3f}
• Level 1: {actual_weights[1]:.3f}  
• Level 2: {actual_weights[2]:.3f}

Attention Statistics:
• Max Attention: {multilevel_fusion.max():.4f}
• Mean Attention: {multilevel_fusion.mean():.4f}
• Std Attention: {multilevel_fusion.std():.4f}

Core Coverage:
• Core Pixels: {np.sum(core_mask > 0):,}
• High Attention: {np.sum(high_attention_masked):,}
• Coverage: {100*np.sum(high_attention_masked)/np.sum(core_mask > 0):.1f}%"""

        # Add metrics summary if available
        if level_metrics:
            stats_text += f"\n\nPerformance Metrics (Best):"
            for level_name in ['Level_0', 'Level_1', 'Level_2', 'Multilevel_Fusion']:
                if level_name in level_metrics:
                    best_f1 = max([m['f1_score'] for m in level_metrics[level_name].values()])
                    best_iou = max([m['iou'] for m in level_metrics[level_name].values()])
                    display_name = level_name.replace('_', ' ')
                    stats_text += f"\n• {display_name}: F1={best_f1:.3f}, IoU={best_iou:.3f}"
        
        plt.text(0.05, 0.95, stats_text, transform=plt.gca().transAxes, 
                fontsize=9, verticalalignment='top', fontfamily='monospace',
                bbox=dict(boxstyle='round', facecolor='lightgray', alpha=0.8))

        # Empty spaces for alignment
        plt.subplot(3, 6, 17)
        plt.axis('off')
        plt.subplot(3, 6, 18)
        plt.axis('off')

        # Add main title
        plt.suptitle(f'HierGAT Comprehensive Analysis: {Path(image_path).stem}', 
                    fontsize=16, fontweight='bold', y=0.98)

        # Adjust layout
        plt.tight_layout(rect=[0, 0, 1, 0.96])

        if save_dir:
            save_path = save_dir / f'hiergat_comprehensive_{Path(image_path).stem}.png'
            plt.savefig(save_path, bbox_inches='tight', dpi=300)
            print(f"✅ Saved HierGAT visualization to: {save_path}")

        plt.close()

        # Return results including metrics
        return {
            'multilevel_fusion': multilevel_fusion,
            'level_maps': smooth_maps,
            'fusion_weights': actual_weights,
            'core_mask': core_mask,
            'level_metrics': level_metrics,
            'statistics': {
                'max_attention': float(multilevel_fusion.max()),
                'mean_attention': float(multilevel_fusion.mean()),
                'std_attention': float(multilevel_fusion.std()),
                'core_pixels': int(np.sum(core_mask > 0)),
                'high_attention_pixels': int(np.sum(high_attention_masked))
            }
        }

    except Exception as e:
        print(f"Error in HierGAT visualization for {Path(image_path).stem}: {str(e)}")
        plt.close()
        traceback.print_exc()
        return None


def main():
    """
    Main function for HierGAT-specific visualization.
    """
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='HierGAT Comprehensive Visualization')
    parser.add_argument('--model_path', type=str, 
                       default="../../training_step_2/self_supervised_training/output/best_model.pt",
                       help='Path to the HierGAT model')
    parser.add_argument('--mil_model_path', type=str,
                       default="../../training_step_1/mil_classification/output/best_model.pth", 
                       help='Path to the MIL model')
    parser.add_argument('--dataset_dir', type=str,
                       default="../../dataset/training_dataset_step_1/tma_core",
                       help='Directory containing images to process')
    parser.add_argument('--save_dir', type=str,
                       default="./output/hiergat_visualization_results",
                       help='Directory to save visualizations')
    parser.add_argument('--mask_dir', type=str,
                       default="../../dataset/training_dataset_step_2/mask",
                       help='Directory containing annotation masks')
    parser.add_argument('--level_weights', type=float, nargs=3, default=[0.5, 0.3, 0.2],
                       help='Weights for Level 0, Level 1, Level 2 (default: 0.5 0.3 0.2)')
    parser.add_argument('--single_image', type=str,
                       help='Process only this specific image')
    parser.add_argument('--output_suffix', type=str,
                       help='Add suffix to output directory')
    parser.add_argument('--calculate_metrics', action='store_true', default=True,
                       help='Calculate performance metrics (default: True)')
    parser.add_argument('--metrics_thresholds', type=float, nargs='+', 
                       default=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
                       help='Thresholds for metrics calculation (default: 0.1 to 0.9 in 0.1 steps)')
    
    args = parser.parse_args()
    
    # Validate level weights
    if len(args.level_weights) != 3:
        raise ValueError("level_weights must contain exactly 3 values for Level 0, 1, 2")
    
    if any(w < 0 for w in args.level_weights):
        raise ValueError("All level weights must be non-negative")
    
    # Normalize weights if they don't sum to 1
    weight_sum = sum(args.level_weights)
    if abs(weight_sum - 1.0) > 1e-6:
        args.level_weights = [w/weight_sum for w in args.level_weights]
        print(f"Normalized weights to: {args.level_weights}")
    
    # Set up paths
    model_path = args.model_path
    mil_model_path = args.mil_model_path
    dataset_dir = args.dataset_dir
    save_dir = args.save_dir
    mask_dir = args.mask_dir
    
    # Add suffix to save directory if provided
    if args.output_suffix:
        save_dir = f"{save_dir}_{args.output_suffix}"
    
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Using device: {device}")
    print(f"Level weights: Level0={args.level_weights[0]:.3f}, Level1={args.level_weights[1]:.3f}, Level2={args.level_weights[2]:.3f}")
    print(f"Calculate metrics: {args.calculate_metrics}")
    if args.calculate_metrics:
        print(f"Metrics thresholds: {args.metrics_thresholds}")
    
    # Initialize HierGAT inference
    try:
        gat_inference = HierGATSSLInference(model_path, mil_model_path=mil_model_path)
        print(f"✅ HierGAT model loaded successfully")
    except Exception as e:
        print(f"❌ Error loading HierGAT model: {str(e)}")
        raise
    
    # Determine images to process
    if args.single_image:
        if Path(args.single_image).exists():
            image_paths = [Path(args.single_image)]
            print(f"Processing single image: {args.single_image}")
        else:
            print(f"❌ Image not found: {args.single_image}")
            return 1
    else:
        image_paths = list(Path(dataset_dir).glob('*.png'))
        print(f"Found {len(image_paths)} images to process")
    
    if not image_paths:
        print("❌ No images found to process")
        return 1
    
    # Process images
    results_summary = []
    
    for image_path in image_paths:
        try:
            print(f"\n{'='*50}")
            print(f"Processing: {image_path.name}")
            print(f"{'='*50}")
            
            # Get HierGAT results
            gat_results = gat_inference.process_image(str(image_path))
            
            # Generate comprehensive visualization
            viz_results = hiergat_comprehensive_visualization(
                gat_results, 
                str(image_path), 
                save_dir=save_dir,
                mask_dir=mask_dir,
                level_weights=args.level_weights,
                calculate_metrics=args.calculate_metrics,
                metrics_thresholds=args.metrics_thresholds
            )
            
            if viz_results is not None:
                print(f"✅ Successfully processed {image_path.name}")
                print(f"   Fusion weights used: {viz_results['fusion_weights']}")
                print(f"   Max attention: {viz_results['statistics']['max_attention']:.4f}")
                print(f"   Coverage: {100*viz_results['statistics']['high_attention_pixels']/viz_results['statistics']['core_pixels']:.1f}%")
                
                result_entry = {
                    'image': image_path.name,
                    'status': 'success',
                    'statistics': viz_results['statistics'],
                    'weights': viz_results['fusion_weights'].tolist()
                }
                
                # Add metrics summary if available
                if 'level_metrics' in viz_results and viz_results['level_metrics']:
                    metrics_summary = {}
                    for level_name, level_data in viz_results['level_metrics'].items():
                        # Get best metrics across all thresholds
                        best_f1 = max([m['f1_score'] for m in level_data.values()])
                        best_iou = max([m['iou'] for m in level_data.values()])
                        best_precision = max([m['precision'] for m in level_data.values()])
                        best_recall = max([m['recall'] for m in level_data.values()])
                        
                        metrics_summary[level_name] = {
                            'best_f1': best_f1,
                            'best_iou': best_iou,
                            'best_precision': best_precision,
                            'best_recall': best_recall
                        }
                        
                        print(f"   📊 {level_name.replace('_', ' ')}: F1={best_f1:.3f}, IoU={best_iou:.3f}")
                    
                    result_entry['metrics_summary'] = metrics_summary
                
                results_summary.append(result_entry)
                
            else:
                print(f"⚠️ Failed to process {image_path.name}")
                results_summary.append({
                    'image': image_path.name,
                    'status': 'failed'
                })
                
        except Exception as e:
            print(f"❌ Error processing {image_path.name}: {str(e)}")
            results_summary.append({
                'image': image_path.name,
                'status': 'error',
                'error': str(e)
            })
            continue

    # Save results summary
    if save_dir:
        summary_path = Path(save_dir) / 'processing_summary.json'
        
        # Calculate overall metrics statistics
        overall_metrics = {}
        successful_results = [r for r in results_summary if r['status'] == 'success' and 'metrics_summary' in r]
        
        if successful_results:
            for level_name in ['Level_0', 'Level_1', 'Level_2', 'Multilevel_Fusion']:
                level_metrics = [r['metrics_summary'][level_name] for r in successful_results if level_name in r['metrics_summary']]
                
                if level_metrics:
                    overall_metrics[level_name] = {
                        'avg_f1': np.mean([m['best_f1'] for m in level_metrics]),
                        'std_f1': np.std([m['best_f1'] for m in level_metrics]),
                        'avg_iou': np.mean([m['best_iou'] for m in level_metrics]),
                        'std_iou': np.std([m['best_iou'] for m in level_metrics]),
                        'avg_precision': np.mean([m['best_precision'] for m in level_metrics]),
                        'std_precision': np.std([m['best_precision'] for m in level_metrics]),
                        'avg_recall': np.mean([m['best_recall'] for m in level_metrics]),
                        'std_recall': np.std([m['best_recall'] for m in level_metrics]),
                        'num_samples': len(level_metrics)
                    }
        
        summary_data = {
            'level_weights': args.level_weights,
            'total_images': len(image_paths),
            'successful': len([r for r in results_summary if r['status'] == 'success']),
            'failed': len([r for r in results_summary if r['status'] != 'success']),
            'with_metrics': len(successful_results),
            'overall_metrics': overall_metrics,
            'results': results_summary
        }
        
        with open(summary_path, 'w') as f:
            json.dump(summary_data, f, indent=2)
        print(f"\n📊 Results summary saved to: {summary_path}")
        
        # Create overall metrics Excel file
        if overall_metrics:
            overall_excel_path = Path(save_dir) / 'overall_metrics_summary.xlsx'
            with pd.ExcelWriter(overall_excel_path, engine='openpyxl') as writer:
                
                # Create summary DataFrame
                summary_data_list = []
                for level_name, metrics in overall_metrics.items():
                    summary_data_list.append({
                        'Level': level_name.replace('_', ' '),
                        'Avg_F1': round(metrics['avg_f1'], 4),
                        'Std_F1': round(metrics['std_f1'], 4),
                        'Avg_IoU': round(metrics['avg_iou'], 4),
                        'Std_IoU': round(metrics['std_iou'], 4),
                        'Avg_Precision': round(metrics['avg_precision'], 4),
                        'Std_Precision': round(metrics['std_precision'], 4),
                        'Avg_Recall': round(metrics['avg_recall'], 4),
                        'Std_Recall': round(metrics['std_recall'], 4),
                        'Num_Samples': metrics['num_samples']
                    })
                
                df_summary = pd.DataFrame(summary_data_list)
                df_summary.to_excel(writer, sheet_name='Overall_Summary', index=False)
                
                # Format the sheet
                workbook = writer.book
                worksheet = writer.sheets['Overall_Summary']
                
                # Format headers
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 15)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📊 Overall metrics summary Excel saved to: {overall_excel_path}")

    print(f"\n🎉 HierGAT visualization processing completed!")
    print(f"📊 Total: {len(image_paths)}, Success: {len([r for r in results_summary if r['status'] == 'success'])}, Failed: {len([r for r in results_summary if r['status'] != 'success'])}")
    
    # Print overall metrics summary
    if successful_results:
        print(f"📈 Images with metrics: {len(successful_results)}")
        print(f"\n📊 Overall Performance Summary:")
        for level_name, metrics in overall_metrics.items():
            print(f"   {level_name.replace('_', ' ')}:")
            print(f"      F1: {metrics['avg_f1']:.3f} ± {metrics['std_f1']:.3f}")
            print(f"      IoU: {metrics['avg_iou']:.3f} ± {metrics['std_iou']:.3f}")
            print(f"      Samples: {metrics['num_samples']}")
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main()) 