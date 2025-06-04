from utils.imports import *
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
import math
import cv2
from pytorch_grad_cam import GradCAM, HiResCAM, ScoreCAM, GradCAMPlusPlus, AblationCAM, XGradCAM, EigenCAM, FullGrad
from pytorch_grad_cam.utils.model_targets import ClassifierOutputTarget
from data.slide_processor import CustomSlide
from utils.extract_core_mask import CoreExtractor
from utils.model_utils import MILModelWrapper, get_target_layer_for_cam, get_model_transform
from .cam_visualizer import create_cam_visualizer
import torch
import torch.nn as nn
import numpy as np
import cv2
from pathlib import Path
import matplotlib.pyplot as plt
from scipy.ndimage import gaussian_filter
from PIL import Image
import traceback
import timm
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


class ModelWrapper(nn.Module):
    """Wrapper for MIL model to make it compatible with CAM methods."""
    
    def __init__(self, model):
        super(ModelWrapper, self).__init__()
        self.model = model
    
    def forward(self, x):
        # CAM methods expect direct input, add batch dimension for MIL model
        x = x.unsqueeze(1)
        _, _, logits, _ = self.model(x)
        return logits


def _get_target_layer(model):
    """Get the target layer for CAM visualization."""
    return get_target_layer_for_cam(model)


def optimal_attention_fusion(smooth_maps, mil_smooth, cam_smooth):
    """
    Optimal fusion of attention maps combining multiple strategies
    """
    # Normalize and smooth GAT maps
    gat_combined = 0.0
    for level, sigma in [(0,1), (1,2), (2,4)]:
        weight = 0.5 if level == 0 else 0.3 if level == 1 else 0.2
        level_map = smooth_maps[level]
        # Normalize before smoothing
        level_map = (level_map - level_map.min()) / (level_map.max() - level_map.min() + 1e-8)
        gat_combined += weight * gaussian_filter(level_map, sigma=sigma)
    
    # Normalize final GAT combination
    gat_combined = (gat_combined - gat_combined.min()) / (gat_combined.max() - gat_combined.min() + 1e-8)
    
    # Keep already normalized and smoothed maps
    mil_enhanced = mil_smooth.copy()
    cam_enhanced = cam_smooth.copy()
    
    # Calculate confidence scores
    gat_conf = np.mean(gat_combined[gat_combined > np.percentile(gat_combined, 90)])
    mil_conf = np.mean(mil_enhanced[mil_enhanced > np.percentile(mil_enhanced, 90)])
    cam_conf = np.mean(cam_enhanced[cam_enhanced > np.percentile(cam_enhanced, 90)])
    
    # Adaptive weights based on confidence
    total_conf = gat_conf + mil_conf + cam_conf
    weights = {
        'gat': 0.2 * (gat_conf / total_conf),
        'mil': 0.5 * (mil_conf / total_conf),
        'fullgrad': 0.3 * (cam_conf / total_conf)  # Keep 'fullgrad' key for compatibility
    }
    
    # Final combination
    combined = (
        weights['gat'] * gat_combined +
        weights['mil'] * mil_enhanced +
        weights['fullgrad'] * cam_enhanced
    )
    
    # Normalize final result
    combined = (combined - combined.min()) / (combined.max() - combined.min() + 1e-8)
    
    return combined, weights, {
        'gat_combined': gat_combined,
        'mil_enhanced': mil_enhanced,
        'fullgrad_enhanced': cam_enhanced  # Keep 'fullgrad_enhanced' key for compatibility
    }


def combine_all_attention_maps(gat_results, mil_model, image_path, device, save_dir=None, mask_dir=None, cam_method='fullgrad', fusion_method='confidence', calculate_metrics=True, metrics_thresholds=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]):
   """
   Combine all attention maps with deterministic GAT and MIL processing
   
   Args:
       gat_results: Results from GAT inference
       mil_model: MIL model for attention extraction
       image_path: Path to the image
       device: Torch device
       save_dir: Directory to save visualization
       mask_dir: Directory containing mask files
       cam_method: CAM method to use ('fullgrad', 'gradcam', 'hirescam', etc.)
       fusion_method: Fusion method for final heatmap ('optimal', 'weighted', 'adaptive', 'multiscale', 'confidence')
       calculate_metrics: Whether to calculate performance metrics
       metrics_thresholds: List of thresholds for metrics calculation
   """
   try:
       if save_dir:
           save_dir = Path(save_dir)
           save_dir.mkdir(parents=True, exist_ok=True)
           
       # Initialize core extractor
       core_extractor = CoreExtractor(threshold=200, kernel_size=7)
           
       # Read original image and get core mask
       original = cv2.imread(str(image_path))
       original = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
       H, W = original.shape[:2]
       
       # Extract core mask and contours
       core_mask = core_extractor.extract_core_mask(original)
       core_contours = core_extractor.get_core_boundary(core_mask)
       
       # Initialize attention maps
       gat_maps = {level: np.zeros((H, W)) for level in range(3)}
       mil_map = np.zeros((H, W))
       cam_map = np.zeros((H, W))
       
       # Get data transform - consistent with step_1_part_2
       transform = get_model_transform(mil_model)
       
       # Process GAT levels deterministically
       viz_data = gat_results['viz_data']
       for level in range(3):
           level_patches = [p for p in viz_data['patches']['kept'] if p['level'] == level]
           level_patches.sort(key=lambda x: (x['level_coords'][1], x['level_coords'][0]))
           
           if not level_patches:
               continue
               
           level_attention = gat_results['attention_weights'][f'level_{level}'].cpu().numpy().squeeze()
           level_attention = (level_attention - level_attention.min()) / (level_attention.max() - level_attention.min() + 1e-8)
           
           scale_factor = 2 ** level
           for idx, patch in enumerate(level_patches):
               if idx >= len(level_attention):
                   break
                   
               base_x, base_y = patch['base_coords']
               patch_size = patch['size'][0] * scale_factor
               
               y_end = min(base_y + patch_size, H)
               x_end = min(base_x + patch_size, W)
               
               if y_end > base_y and x_end > base_x:
                   gat_maps[level][base_y:y_end, base_x:x_end] = level_attention[idx]
       
       # Initialize CAM visualizer - consistent with step_1_part_2
       cam_visualizer = create_cam_visualizer(mil_model, device, cam_method=cam_method, target_class=1)
       cam_method_name = cam_visualizer.get_method_name()
       
       # Process level 0 patches deterministically for MIL and CAM
       level0_patches = [p for p in viz_data['patches']['kept'] if p['level'] == 0]
       level0_patches.sort(key=lambda x: (x['level_coords'][1], x['level_coords'][0]))
       patch_size = 224
       
       valid_patches = []
       valid_coords = []
       patches_data = []  # For CAM processing
       
       for patch in level0_patches:
           x, y = patch['level_coords']
           
           if y + patch_size > H or x + patch_size > W:
               continue
               
           try:
               patch_img = original[y:y+patch_size, x:x+patch_size]
               if patch_img.shape[:2] != (patch_size, patch_size):
                   continue
                   
               patch_pil = Image.fromarray(patch_img)
               patch_tensor = transform(patch_pil).unsqueeze(0).to(device)
               
               # Process CAM using the new visualizer - consistent with step_1_part_2
               grayscale_cam = cam_visualizer.process_patch(patch_img, (x, y))
               
               if grayscale_cam is not None:
                   valid_patches.append(patch_tensor)
                   valid_coords.append((x, y))
                   
                   y_end = min(y + patch_size, H)
                   x_end = min(x + patch_size, W)
                   cam_map[y:y_end, x:x_end] = grayscale_cam
               
           except Exception as e:
               print(f"Error processing patch at ({x}, {y}): {str(e)}")
               continue
       
       # Process MIL attention deterministically - consistent with step_1_part_2
       if valid_patches:
           patch_batch = torch.cat(valid_patches, dim=0)
           with torch.no_grad():
               _, _, _, mil_attention = mil_model(patch_batch.unsqueeze(0))
           
           mil_attention = mil_attention.cpu().numpy().squeeze()
           mil_attention = (mil_attention - mil_attention.min()) / (mil_attention.max() - mil_attention.min() + 1e-8)
           
           for idx, (x, y) in enumerate(valid_coords):
               if idx >= len(mil_attention):
                   break
                   
               y_end = min(y + patch_size, H)
               x_end = min(x + patch_size, W)
               mil_map[y:y_end, x:x_end] = mil_attention[idx]
       
       # Apply core masks
       for level in gat_maps:
           gat_maps[level] = core_extractor.apply_core_mask(gat_maps[level], core_mask)
       mil_map = core_extractor.apply_core_mask(mil_map, core_mask)
       cam_map = core_extractor.apply_core_mask(cam_map, core_mask)
       
       # Apply normalization and smoothing - consistent with step_1_part_2
       smooth_maps = {}
       for level in range(3):
           # First normalize
           level_map = gat_maps[level].copy()
           level_map = (level_map - level_map.min()) / (level_map.max() - level_map.min() + 1e-8)
          
           # Then apply smoothing
           sigma = 150 * (2 ** level)
           smoothed = gaussian_filter(level_map, sigma=sigma)
          
           # Apply core mask and store
           smooth_maps[level] = core_extractor.apply_core_mask(smoothed, core_mask)

       # Normalize MIL map before smoothing
       mil_map = (mil_map - mil_map.min()) / (mil_map.max() - mil_map.min() + 1e-8)
       mil_smooth = gaussian_filter(mil_map, sigma=150)
       mil_smooth = core_extractor.apply_core_mask(mil_smooth, core_mask)

       # Normalize CAM map before smoothing
       cam_map = (cam_map - cam_map.min()) / (cam_map.max() - cam_map.min() + 1e-8)
       cam_smooth = gaussian_filter(cam_map, sigma=150)
       cam_smooth = core_extractor.apply_core_mask(cam_smooth, core_mask)

       # Apply configurable attention fusion
       print(f"🔀 Applying {fusion_method} fusion method...")
       
       # First create the enhanced_maps using optimal fusion for individual components
       combined_attention, fusion_weights, enhanced_maps = optimal_attention_fusion(
           smooth_maps, mil_smooth, cam_smooth
       )
       
       # Initialize attention fusion for final fusion heatmap
       attention_fusion = AttentionFusion()
       
       # Prepare attention maps for final fusion (the three components)
       final_fusion_maps = {
           'multilevel_fused': enhanced_maps['gat_combined'],  # Multilevel fused heatmap 
           'mil': enhanced_maps['mil_enhanced'],              # MIL attention
           'cam': enhanced_maps['fullgrad_enhanced']          # CAM based
       }
       
       # Apply final fusion using the specified method
       final_combination = ('multilevel_fused', 'mil', 'cam')
       fusion_params = {
           'weights': {
               '3type': {'gat': 0.4, 'mil': 0.4, 'grad': 0.2}  # Balanced weights for final fusion
           },
           'base_sigma': 150
       }
       
       final_fusion_result = attention_fusion.apply_fusion(
           final_fusion_maps,
           final_combination, 
           fusion_type=fusion_method,
           fusion_params=fusion_params
       )
       
       # Get final fused heatmap
       final_fusion_heatmap = final_fusion_result['fused_map']
       final_fusion_weights = final_fusion_result.get('weights', {})
       
       print(f"✅ Final fusion completed using {fusion_method} method")
       print(f"Final fusion weights: {final_fusion_weights}")
       
       # Add final fusion heatmap to enhanced_maps for metrics
       enhanced_maps['final_fusion'] = final_fusion_heatmap
       
       # Load corresponding annotated mask
       image_name = Path(image_path).name
       if mask_dir:
           mask_path = Path(mask_dir) / image_name
       else:
           mask_path = Path("/scratch/nk53/rm8989/gene_prediction/code/GRAPHITE/mask_resized") / image_name
      
       # Initialize metrics variables
       fusion_metrics = None
       
       if mask_path.exists():
           annotated_mask = cv2.imread(str(mask_path))
           annotated_mask = cv2.cvtColor(annotated_mask, cv2.COLOR_BGR2RGB)
           
           # Calculate performance metrics for all fusion methods including final fusion
           if calculate_metrics:
               print(f"📊 Calculating performance metrics for {image_name}...")
               fusion_metrics = calculate_fusion_metrics(
                   enhanced_maps, annotated_mask, cam_method=cam_method, thresholds=metrics_thresholds
               )
               
               # Create Excel file with metrics
               if save_dir and fusion_metrics:
                   excel_path = create_fusion_metrics_excel(
                       image_name, fusion_metrics, save_dir, cam_method=cam_method
                   )
           else:
               print(f"📊 Skipping metrics calculation for {image_name}")
           
       else:
           print(f"Warning: Annotated mask not found for {image_name}")
           annotated_mask = np.zeros_like(original)

       # Visualization code
       plt.figure(figsize=(15, 15))

       # First Row - Actual Tissue Core in middle
       plt.subplot(431)
       plt.axis('off')
       plt.subplot(432)
       plt.imshow(original)
       for contour in core_contours:
           plt.plot(contour[:, 0, 0], contour[:, 0, 1], 'r-', linewidth=2)
       plt.title('Actual Tissue Core', fontsize=12)
       plt.axis('off')
       plt.subplot(433)
       plt.axis('off')

       # Second Row - GAT Levels (0-2)
       for level in range(3):
           plt.subplot(4, 3, level + 4)
           plt.imshow(smooth_maps[level], cmap='jet', vmin=0, vmax=1)
           plt.title(f'Level {level}', fontsize=12)
           plt.axis('off')

       # Third Row - CAM, MIL, Multi-Level Fusion
       plt.subplot(437)
       plt.imshow(enhanced_maps['fullgrad_enhanced'], cmap='jet', vmin=0, vmax=1)
       plt.title(cam_method_name, fontsize=12)
       plt.axis('off')

       plt.subplot(438)
       plt.imshow(enhanced_maps['mil_enhanced'], cmap='jet', vmin=0, vmax=1)
       plt.title('MIL Attention', fontsize=12)
       plt.axis('off')

       plt.subplot(439)
       plt.imshow(enhanced_maps['gat_combined'], cmap='jet', vmin=0, vmax=1)
       plt.title('Multi-Level Fusion', fontsize=12)
       plt.axis('off')

       # Fourth Row - Final Fusion, Final Overlay, Annotated Mask
       plt.subplot(4,3,10)
       plt.imshow(final_fusion_heatmap, cmap='jet', vmin=0, vmax=1)
       plt.title(f'Final Fusion ({fusion_method.title()})', fontsize=12)
       plt.axis('off')

       # Final Overlay with colorbar
       plt.subplot(4,3,11)
       attention_threshold = 0.1
       plt.imshow(original)
       attention_mask = np.ma.masked_array(
           final_fusion_heatmap,
           mask=~((final_fusion_heatmap > attention_threshold) & (core_mask > 0))
       )
       im = plt.imshow(attention_mask, cmap='jet', alpha=0.5, vmin=0, vmax=1)
       plt.colorbar(im, orientation='vertical')
       plt.title('Final Overlay', fontsize=12)
       plt.axis('off')

       # Annotated Mask instead of Tissue Mask
       plt.subplot(4,3,12)
       plt.imshow(annotated_mask)
       plt.title('Annotated Mask', fontsize=12)
       plt.axis('off')

       # Adjust layout
       plt.tight_layout(rect=[0, 0, 1, 0.95])

       if save_dir:
           save_path = save_dir / f'multiattention_{cam_method}_{fusion_method}_{Path(image_path).stem}.png'
           plt.savefig(save_path, bbox_inches='tight', dpi=300)
           print(f"Saved visualization to: {save_path}")

       plt.close()

       return final_fusion_heatmap, smooth_maps, final_fusion_weights, fusion_metrics

   except Exception as e:
       print(f"Error in visualization for {Path(image_path).stem}: {str(e)}")
       plt.close()
       traceback.print_exc()
       return None, None, None, None

def save_visualization_data(output_dir, kept_patches, patch_metadata, slide):
    """
    Save enhanced visualization data including level-specific thresholds and patch information
    """
    viz_data = {
        'image_dimensions': {
            'width': slide.original_dimensions[0],
            'height': slide.original_dimensions[1]
        },
        'levels': slide.levels,
        'level_dimensions': slide.level_dimensions,
        'downsamples': slide.level_downsamples,
        'patch_size': patch_metadata['extraction_params']['patch_size'],
        'level_thresholds': patch_metadata['extraction_params']['level_thresholds'],
        'default_threshold': patch_metadata['extraction_params']['default_threshold'],
        'patches': {
            'kept': [{
                'level': p['level'],
                'level_coords': p['level_coords'],
                'base_coords': p['base_coords'],
                'size': p['size'],
                'downsample': p['downsample'],
                'is_regular': p.get('is_regular', True),
                'edge_type': p.get('edge_type', 'regular')
            } for p in kept_patches],
            'filtered': []
        }
    }
    
    # Add filtered patches with threshold information
    for level in range(slide.levels):
        level_filtered = patch_metadata['filtered_patches'].get(f'level_{level}', [])
        # Add threshold information to each filtered patch
        for patch in level_filtered:
            patch['threshold_used'] = viz_data['level_thresholds'].get(
                level, viz_data['default_threshold']
            )
        viz_data['patches']['filtered'].extend(level_filtered)
    
    viz_path = os.path.join(output_dir, 'visualization_data.json')
    with open(viz_path, 'w') as f:
        json.dump(viz_data, f, indent=2)
    
    return viz_data

def calculate_comprehensive_metrics(pred_attention, gt_mask, thresholds=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]):
    """
    Calculate comprehensive metrics for attention maps vs ground truth masks.
    
    Args:
        pred_attention: Predicted attention map (continuous values 0-1)
        gt_mask: Ground truth binary mask
        thresholds: List of thresholds to convert attention to binary mask
        
    Returns:
        Dictionary containing metrics for each threshold
    """
    # Normalize ground truth to binary
    gt_binary = (gt_mask > 0).astype(int)
    
    # Convert RGB to grayscale if needed
    if len(gt_binary.shape) == 3:
        gt_binary = np.mean(gt_binary, axis=2)
        gt_binary = (gt_binary > 0.5).astype(int)
    
    results = {}
    
    for threshold in thresholds:
        # Convert attention to binary mask
        pred_binary = (pred_attention > threshold).astype(int)
        
        # Calculate confusion matrix elements
        TP = np.sum((pred_binary == 1) & (gt_binary == 1))
        FP = np.sum((pred_binary == 1) & (gt_binary == 0))
        FN = np.sum((pred_binary == 0) & (gt_binary == 1))
        TN = np.sum((pred_binary == 0) & (gt_binary == 0))
        
        # Calculate metrics
        precision = TP / (TP + FP) if (TP + FP) > 0 else 0
        recall = TP / (TP + FN) if (TP + FN) > 0 else 0
        f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        iou = TP / (TP + FP + FN) if (TP + FP + FN) > 0 else 0
        accuracy = (TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) > 0 else 0
        specificity = TN / (TN + FP) if (TN + FP) > 0 else 0
        
        results[f'threshold_{threshold}'] = {
            'threshold': threshold,
            'precision': precision,
            'recall': recall,
            'f1_score': f1_score,
            'iou': iou,
            'accuracy': accuracy,
            'specificity': specificity,
            'TP': TP,
            'FP': FP,
            'FN': FN,
            'TN': TN
        }
    
    return results


def create_fusion_metrics_excel(image_name, fusion_metrics, save_dir, cam_method='fullgrad'):
    """
    Create Excel file with metrics for all three fusion methods.
    
    Args:
        image_name: Name of the image/core
        fusion_metrics: Dictionary containing metrics for each fusion method
        save_dir: Directory to save the Excel file
        cam_method: CAM method used (for naming)
    """
    try:
        save_dir = Path(save_dir)
        save_dir.mkdir(parents=True, exist_ok=True)
        
        # Create Excel file
        excel_path = save_dir / f'fusion_metrics_{Path(image_name).stem}_{cam_method}.xlsx'
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
            # Define the fusion methods and their corresponding sheet names
            fusion_methods = {
                'Multilevel_Fusion': 'Multilevel Fusion',
                'MIL_Attention': 'MIL Attention', 
                'CAM_Based': f'CAM Based ({cam_method.upper()})',
                'Final_Fusion_Heatmap': 'Final Fusion Heatmap'
            }
            
            for method_key, sheet_name in fusion_methods.items():
                if method_key in fusion_metrics:
                    metrics_data = fusion_metrics[method_key]
                    
                    # Convert to DataFrame
                    df_data = []
                    for threshold_key, metrics in metrics_data.items():
                        df_data.append(metrics)
                    
                    df = pd.DataFrame(df_data)
                    
                    # Round numerical values for better display
                    numeric_cols = ['precision', 'recall', 'f1_score', 'iou', 'accuracy', 'specificity']
                    for col in numeric_cols:
                        if col in df.columns:
                            df[col] = df[col].round(4)
                    
                    # Write to Excel sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Format headers
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 15)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Highlight best performing thresholds
                    if len(df) > 0:
                        # Find best F1 score row
                        best_f1_idx = df['f1_score'].idxmax() + 2  # +2 for 0-indexing and header
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=best_f1_idx, column=col_idx)
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        
                        # Find best IoU score row
                        best_iou_idx = df['iou'].idxmax() + 2
                        if best_iou_idx != best_f1_idx:
                            for col_idx in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=best_iou_idx, column=col_idx)
                                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        
        print(f"✅ Fusion metrics Excel file saved: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Error creating fusion metrics Excel file: {str(e)}")
        traceback.print_exc()
        return None


def calculate_fusion_metrics(enhanced_maps, annotated_mask, cam_method='fullgrad', thresholds=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]):
    """
    Calculate performance metrics for all fusion methods including final fusion.
    
    Args:
        enhanced_maps: Dictionary containing the attention maps
        annotated_mask: Ground truth annotation mask
        cam_method: CAM method used
        thresholds: List of thresholds for metrics calculation
    
    Returns:
        Dictionary containing metrics for each fusion method
    """
    fusion_metrics = {}
    
    # Map enhanced_maps keys to our fusion method names
    method_mapping = {
        'gat_combined': 'Multilevel_Fusion',
        'mil_enhanced': 'MIL_Attention',
        'fullgrad_enhanced': 'CAM_Based',
        'final_fusion': 'Final_Fusion_Heatmap'  # Add final fusion metrics
    }
    
    for map_key, method_name in method_mapping.items():
        if map_key in enhanced_maps:
            print(f"   📈 Calculating metrics for {method_name.replace('_', ' ')}")
            fusion_metrics[method_name] = calculate_comprehensive_metrics(
                enhanced_maps[map_key], annotated_mask, thresholds=thresholds
            )
    
    return fusion_metrics

class AttentionFusion:
    """
    Simple attention fusion class to replace the missing attention.fusion_methods.AttentionFusion
    """
    def __init__(self):
        pass
    
    def apply_fusion(self, attention_maps, combination, fusion_type='confidence', fusion_params=None):
        """
        Apply fusion to attention maps
        
        Args:
            attention_maps: Dictionary of attention maps
            combination: Tuple of map names to combine
            fusion_type: Type of fusion to apply
            fusion_params: Parameters for fusion
        
        Returns:
            Dictionary with 'fused_map' and 'weights'
        """
        if fusion_params is None:
            fusion_params = {}
            
        # Get weights from fusion_params
        weights_config = fusion_params.get('weights', {})
        method_weights = weights_config.get('3type', {'gat': 0.4, 'mil': 0.4, 'grad': 0.2})
        
        # Map combination names to weight keys
        weight_mapping = {
            'multilevel_fused': 'gat',
            'mil': 'mil', 
            'cam': 'grad'
        }
        
        # Combine maps using weighted average
        fused_map = np.zeros_like(list(attention_maps.values())[0])
        total_weight = 0
        used_weights = {}
        
        for map_name in combination:
            if map_name in attention_maps:
                weight_key = weight_mapping.get(map_name, map_name)
                weight = method_weights.get(weight_key, 1.0 / len(combination))
                
                fused_map += weight * attention_maps[map_name]
                total_weight += weight
                used_weights[map_name] = weight
        
        # Normalize by total weight
        if total_weight > 0:
            fused_map /= total_weight
            
        # Normalize to [0, 1] range
        fused_map = (fused_map - fused_map.min()) / (fused_map.max() - fused_map.min() + 1e-8)
        
        return {
            'fused_map': fused_map,
            'weights': used_weights
        }